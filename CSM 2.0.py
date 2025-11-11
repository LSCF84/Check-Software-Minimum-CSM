# ---------- PARTE 1: IMPORTS Y CONSTANTES ----------
import os
import sys
import shutil
import subprocess
import threading
import json
import webbrowser
import zipfile
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *

# Intentar importar openpyxl para exportar a Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[AVISO] openpyxl no instalado. Exportación a Excel deshabilitada.")

APP_NAME = "CSM v.2.0"
if getattr(sys, 'frozen', False):
    DEFAULT_ROOT = os.path.dirname(sys.executable)
else:
    DEFAULT_ROOT = os.path.dirname(os.path.abspath(__file__))

KEYWORDS = ["setup","standard","installer","win","win64","x64","client","install","rainmeter","ldplayer"]
SPECIAL_PORTABLE_PREFIX = "sdio_x64_"

# Base de datos para tooltips
SOFTWARE_INFO = {
    "ldplayer": "LDPlayer: Emulador Android para PC, ideal para juegos móviles.",
    "bluestacks": "BlueStacks: Emulador Android popular para apps y juegos.",
    "rainmeter": "Rainmeter: Personaliza tu escritorio Windows con skins y widgets.",
    "python": "Python: Lenguaje de programación versátil y poderoso.",
    "vlc": "VLC Media Player: Reproduce casi cualquier formato multimedia.",
    "7zip": "7-Zip: Compresor/descompresor de archivos de código abierto.",
    "notepad++": "Notepad++: Editor de texto avanzado para programadores.",
}

def is_minios_folder(path: str) -> bool:
    norm = os.path.normcase(path)
    target = os.path.normcase(os.path.join(DEFAULT_ROOT, "Sistema", "Minios Software"))
    return target in norm
	# ---------- PARTE 2: CLASE PRINCIPAL Y CONSTRUCTOR ----------
class InstallerApp(tb.Window):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1200x760")
        self.minsize(980, 600)
        self.style.theme_use("darkly")

        self.path_map = {}
        self.path_type = {}
        self.check_vars = {}
        self.current_progress = None
        self.total_tasks = 0
        self.current_task = 0
        self.winget_cache = []

        self._build_ui()
        self.check_tools()

    def _build_ui(self):
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True)

        # Pestañas
        self.tab_installers = ttk.Frame(self.notebook)
        self.tab_updates = ttk.Frame(self.notebook)
        self.tab_repo = ttk.Frame(self.notebook)
        self.tab_optimizador = ttk.Frame(self.notebook)
        self.tab_temas = ttk.Frame(self.notebook)
        self.tab_about = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_installers, text="Instaladores")
        self.notebook.add(self.tab_updates, text="Actualizaciones")
        self.notebook.add(self.tab_repo, text="Repo-Software")
        self.notebook.add(self.tab_optimizador, text="Optimizador")
        self.notebook.add(self.tab_temas, text="Temas")
        self.notebook.add(self.tab_about, text="Acerca de")

        # Construir cada pestaña
        self._build_installers_tab()
        self._build_updates_tab()
        self._build_repo_tab()
        self._build_optimizador_tab()
        self._build_themes_tab()
        self._build_about_tab()
		# ---------- PARTE 3: PESTAÑA INSTALADORES (1/2) ----------
    def _build_installers_tab(self):
        top = ttk.Frame(self.tab_installers); top.pack(fill="x", padx=10, pady=8)
        folder_label = ttk.Label(top, text="Carpeta raíz:")
        folder_label.pack(side="left")
        self.create_tooltip(folder_label, f"Ruta actual: {DEFAULT_ROOT}")

        self.folder_var = tk.StringVar(value=DEFAULT_ROOT)
        folder_entry = ttk.Entry(top, textvariable=self.folder_var)
        folder_entry.pack(side="left", fill="x", expand=True, padx=8)
        ttk.Button(top, text="Examinar…", command=self.select_folder).pack(side="left")
        ttk.Button(top, text="Escanear", command=self.scan_now).pack(side="left", padx=(8,0))

        self.tree = ttk.Treeview(self.tab_installers, show="tree")
        self.tree.heading("#0", text="Carpetas e Instaladores")
        self.tree.pack(fill="both", expand=True, padx=10, pady=(6,8))
        self.tree.bind("<Motion>", self.on_tree_motion)
        self.tree.bind("<Button-1>", self.on_tree_click)

        bottom = ttk.Frame(self.tab_installers); bottom.pack(fill="x", padx=10, pady=(0,10))
        ttk.Button(bottom, text="Seleccionar todo", command=lambda: self.set_all_checks(True)).pack(side="left")
        ttk.Button(bottom, text="Deseleccionar todo", command=lambda: self.set_all_checks(False)).pack(side="left", padx=5)
        self.install_btn = ttk.Button(bottom, text="Instalar seleccionados", command=self.install_selected)
        self.install_btn.pack(side="right", padx=5)
        self.run_btn = ttk.Button(bottom, text="Ejecutar seleccionados (portables/txt)", command=self.run_selected)
        self.run_btn.pack(side="right", padx=5)

        self.status = tk.StringVar(value="Listo.")
        ttk.Label(self.tab_installers, textvariable=self.status, anchor="w").pack(fill="x", side="bottom")
		# ---------- PARTE 4: PESTAÑA INSTALADORES (2/2) ----------
    def scan_now(self):
        folder = self.folder_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Carpeta inválida", "Selecciona una carpeta existente.")
            return

        self.status.set("Escaneando…")
        self.update_idletasks()
        self.path_map.clear()
        self.path_type.clear()
        self.check_vars.clear()
        for iid in list(self.tree.get_children()):
            self.tree.delete(iid)
        self._insert_tree("", folder)
        self.status.set("Escaneo completo.")

    def _insert_tree(self, parent, path):
        try:
            entries = sorted(os.listdir(path), key=lambda s: s.lower())
        except Exception:
            return

        cmd_files = [e for e in entries if os.path.isfile(os.path.join(path,e)) and os.path.splitext(e)[1].lower() in ('.bat', '.cmd')]
        files_to_show = []
        if cmd_files:
            files_to_show = sorted(cmd_files, key=lambda s: s.lower())
        else:
            for e in entries:
                full = os.path.join(path,e)
                if os.path.isfile(full):
                    low = e.lower(); ext = os.path.splitext(e)[1].lower()
                    if ext in ('.zip','.rar','.txt','.key','.reg','.cmd','.bat'):
                        files_to_show.append(e)
                    elif ext in ('.exe','.msi','.msixbundle'):
                        if low.startswith(SPECIAL_PORTABLE_PREFIX) or is_minios_folder(path) or 'portable' in low:
                            files_to_show.append(e)
                        elif any(k in low for k in KEYWORDS):
                            files_to_show.append(e)

        for entry in files_to_show:
            full = os.path.join(path, entry); low = entry.lower(); ext = os.path.splitext(entry)[1].lower()
            ptype = None
            if ext in ('.cmd','.bat'): ptype = "portable"
            elif ext == '.txt': ptype = "txt"
            elif ext == '.key' or 'patch' in low: ptype = "key"
            elif ext in ('.zip','.rar'): ptype = "archive"
            elif ext == '.reg': ptype = "reg"
            elif ext in ('.exe','.msi','.msixbundle'):
                if low.startswith(SPECIAL_PORTABLE_PREFIX) or is_minios_folder(path) or 'portable' in low: ptype = "portable"
                elif any(k in low for k in KEYWORDS): ptype = "installer"
            if ptype:
                display = entry
                if 'sdio_auto' in low and ext == '.bat': display = "SDIO_ejecutable"
                iid = self.tree.insert(parent, "end", text=f"☐ {display}")
                self.path_map[iid] = full; self.path_type[iid] = ptype; self.check_vars[iid] = False

        for entry in entries:
            full = os.path.join(path, entry)
            if os.path.isdir(full) and entry.lower() != "Software Instaladores":
                node = self.tree.insert(parent, "end", text=entry, open=False)
                self._insert_tree(node, full)
				# ---------- PARTE 5: PESTAÑA ACTUALIZACIONES ----------
    def _build_updates_tab(self):
        frame = self.tab_updates

        top = ttk.Frame(frame); top.pack(fill="x", padx=10, pady=6)
        ttk.Label(top, text="Comando:").pack(side="left")
        self.cmd_var = tk.StringVar(value="winget upgrade")
        ttk.Entry(top, textvariable=self.cmd_var).pack(side="left", fill="x", expand=True, padx=8)
        ttk.Button(top, text="winget upgrade", command=lambda: self._set_and_run("winget upgrade")).pack(side="left", padx=2)
        ttk.Button(top, text="winget upgrade --all", command=lambda: self._set_and_run("winget upgrade --all")).pack(side="left", padx=2)
        self.exec_cmd_btn = ttk.Button(top, text="Ejecutar", command=self.run_command_with_progress)
        self.exec_cmd_btn.pack(side="left", padx=6)
        ttk.Button(top, text="Limpiar salida", command=lambda: self.term_text.delete("1.0","end")).pack(side="left", padx=6)

        term_frame = ttk.Frame(frame); term_frame.pack(fill="both", expand=True, padx=10, pady=(0,10))
        self.term_text = tk.Text(term_frame, wrap="word", bg="black", fg="lime", insertbackground="lime", font=("Consolas", 10))
        self.term_text.pack(side="left", fill="both", expand=True)
        term_scroll = ttk.Scrollbar(term_frame, orient="vertical", command=self.term_text.yview)
        term_scroll.pack(side="right", fill="y")
        self.term_text.configure(yscrollcommand=term_scroll.set)

    def run_command_with_progress(self):
        cmd = self.cmd_var.get().strip()
        if not cmd:
            return

        self.term_text.delete("1.0", "end")
        self.exec_cmd_btn.config(state="disabled")

        def worker():
            try:
                proc = subprocess.Popen(
                    cmd,
                    shell=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding='utf-8'
                )

                def update_progress(percent):
                    blocks = "█" * (percent // 2) + "▒" * (50 - (percent // 2))
                    self.term_text.insert("end", f"  {blocks}  {percent}%\n")
                    self.term_text.see("end")

                # Simular progreso (winget no da % real en upgrade)
                for i in range(0, 101, 2):
                    self.after(0, lambda p=i: update_progress(p))
                    self.after(50)

                for line in proc.stdout:
                    self.term_text.insert("end", line)
                    self.term_text.see("end")

                rc = proc.wait()
                self.term_text.insert("end", f"[Proceso finalizado con código {rc}]\n\n")
                self.term_text.see("end")
            except Exception as e:
                self.term_text.insert("end", f"[Error] {e}\n\n")
                self.term_text.see("end")
            finally:
                self.after(0, lambda: self.exec_cmd_btn.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()
		# ---------- PARTE 6: PESTAÑA REPO-SOFTWARE ----------
    def _build_repo_tab(self):
        frame = self.tab_repo

        ttk.Label(frame, text="Backup & Restauración con Winget", font=("Segoe UI", 14, "bold")).pack(pady=10)

        btn_frame = ttk.Frame(frame); btn_frame.pack(pady=10)

        ttk.Button(btn_frame, text="Listar Software Instalado", command=self.list_installed_software, width=25).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Generar Backup (.bat + .json + .xlsx + .zip)", command=self.generate_full_backup, width=35).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Abrir winstall.app", command=self.open_winstall, width=20).pack(side="left", padx=5)

        ttk.Label(frame, text="Software detectado (winget list):").pack(anchor="w", padx=10, pady=(10, 0))
        text_frame = ttk.Frame(frame)
        text_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.software_text = tk.Text(text_frame, wrap="none", bg="#2a2a2a", fg="white", font=("Consolas", 10))
        self.software_text.pack(side="left", fill="both", expand=True)

        v_scroll = ttk.Scrollbar(text_frame, orient="vertical", command=self.software_text.yview)
        v_scroll.pack(side="right", fill="y")
        h_scroll = ttk.Scrollbar(frame, orient="horizontal", command=self.software_text.xview)
        h_scroll.pack(fill="x", padx=10)
        self.software_text.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.repo_status = tk.StringVar(value="Listo. Usa 'Listar Software Instalado' para comenzar.")
        ttk.Label(frame, textvariable=self.repo_status, anchor="w").pack(fill="x", padx=10, pady=5)

    def list_installed_software(self):
        self.repo_status.set("Obteniendo lista de software instalado...")
        self.software_text.delete("1.0", "end")
        self.software_text.insert("1.0", "Ejecutando 'winget list'...\n\n")

        def worker():
            try:
                proc = subprocess.Popen(
                    ["winget", "list"],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding='utf-8'
                )
                output = ""
                for line in proc.stdout:
                    output += line
                proc.wait()

                if proc.returncode == 0:
                    self.after(0, lambda: self.software_text.delete("1.0", "end"))
                    self.after(0, lambda: self.software_text.insert("1.0", output))
                    self.after(0, lambda: self.repo_status.set("Software listado correctamente."))
                else:
                    self.after(0, lambda: self.software_text.insert("end", "\n[Error al ejecutar winget list]"))
                    self.after(0, lambda: self.repo_status.set("Error al listar software."))
            except Exception as e:
                self.after(0, lambda: self.software_text.insert("end", f"\n[Excepción: {str(e)}]"))
                self.after(0, lambda: self.repo_status.set(f"Error: {str(e)}"))

        threading.Thread(target=worker, daemon=True).start()

    def extract_package_data(self):
        """Extrae datos completos: Nombre, Id, Versión, Disponible, Origen."""
        content = self.software_text.get("1.0", "end").strip()
        if not content or "Nombre" not in content:
            return None

        lines = content.splitlines()
        packages = []
        header_found = False
        col_positions = {}

        for line in lines:
            if "Nombre" in line and "Id" in line and "Versión" in line:
                header_found = True
                col_positions = {
                    "nombre_start": line.find("Nombre"),
                    "id_start": line.find("Id"),
                    "version_start": line.find("Versión"),
                    "disponible_start": line.find("Disponible"),
                    "origen_start": line.find("Origen")
                }
                if any(pos == -1 for pos in col_positions.values()):
                    continue
                continue

            if not header_found or not line.strip() or line.startswith("-"):
                continue

            nombre = line[col_positions["nombre_start"]:col_positions["id_start"]].strip()
            pkg_id = line[col_positions["id_start"]:col_positions["version_start"]].strip()
            version = line[col_positions["version_start"]:col_positions["disponible_start"]].strip() if col_positions["disponible_start"] != -1 else ""
            disponible = line[col_positions["disponible_start"]:col_positions["origen_start"]].strip() if col_positions["origen_start"] != -1 else ""
            origen = line[col_positions["origen_start"]:].strip() if col_positions["origen_start"] != -1 else ""

            if pkg_id and pkg_id != "Id":
                packages.append({
                    "nombre": nombre,
                    "id": pkg_id,
                    "version": version,
                    "disponible": disponible,
                    "origen": origen
                })

        return packages

    def generate_full_backup(self):
        packages = self.extract_package_data()
        if packages is None:
            messagebox.showwarning("Advertencia", "Primero debes listar el software instalado.")
            return
        if not packages:
            messagebox.showwarning("Advertencia", "No se encontraron paquetes válidos.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"winget-backup-{timestamp}"

        dest_folder = filedialog.askdirectory(title="Selecciona carpeta de destino para el backup")
        if not dest_folder:
            return

        bat_path = os.path.join(dest_folder, f"{base_name}.bat")
        json_path = os.path.join(dest_folder, f"{base_name}.json")
        zip_path = os.path.join(dest_folder, f"{base_name}.zip")
        excel_path = os.path.join(dest_folder, f"{base_name}.xlsx")

        # Generar .bat
        try:
            with open(bat_path, "w", encoding="utf-8") as f:
                f.write("@echo off\n")
                f.write(f"REM Backup generado por CSM v.2.0 el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("REM Compatible con Pckgr / Intune\n")
                f.write("echo Iniciando restauración...\n\n")
                for pkg in packages:
                    f.write(f'winget install -e --id "{pkg["id"]}"\n')
                f.write('\necho ¡Restauración completada!\npause\n')
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el archivo .bat:\n{e}")
            return

        # Generar .json
        try:
            data = {"apps": [{"id": pkg["id"]} for pkg in packages]}
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el archivo .json:\n{e}")
            return

        # Generar Excel
        if EXCEL_AVAILABLE:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Software Instalado"

                headers = ["Nombre", "Id", "Versión", "Disponible", "Origen"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                for row_num, pkg in enumerate(packages, 2):
                    ws.cell(row=row_num, column=1, value=pkg["nombre"])
                    ws.cell(row=row_num, column=2, value=pkg["id"])
                    ws.cell(row=row_num, column=3, value=pkg["version"])
                    ws.cell(row=row_num, column=4, value=pkg["disponible"])
                    ws.cell(row=row_num, column=5, value=pkg["origen"])

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = min(adjusted_width, 50)

                wb.save(excel_path)
            except Exception as e:
                messagebox.showwarning("Advertencia Excel", f"No se pudo crear el archivo Excel:\n{e}")

        # Comprimir en .zip
        try:
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(bat_path, os.path.basename(bat_path))
                zipf.write(json_path, os.path.basename(json_path))
                if EXCEL_AVAILABLE and os.path.exists(excel_path):
                    zipf.write(excel_path, os.path.basename(excel_path))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el archivo .zip:\n{e}")
            return

        msg = f"✅ Archivos generados en:\n{dest_folder}\n\n"
        msg += f" - {os.path.basename(bat_path)}\n"
        msg += f" - {os.path.basename(json_path)}\n"
        if EXCEL_AVAILABLE:
            msg += f" - {os.path.basename(excel_path)}\n"
        msg += f" - {os.path.basename(zip_path)}\n\n"
        msg += "El archivo .json es compatible con Pckgr para despliegues en Intune."

        messagebox.showinfo("Backup Completo", msg)
        self.repo_status.set(f"Backup completo generado: {zip_path}")
		# ---------- PARTE 7: PESTAÑA OPTIMIZADOR ----------
    def _build_optimizador_tab(self):
        frame = self.tab_optimizador

        # Scroll
        canvas = tk.Canvas(frame)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Dos columnas
        columns_frame = ttk.Frame(scrollable_frame)
        columns_frame.pack(fill='both', expand=True, padx=10, pady=10)

        left_col = ttk.Frame(columns_frame)
        left_col.pack(side="left", fill="both", expand=True, padx=5)

        right_col = ttk.Frame(columns_frame)
        right_col.pack(side="right", fill="both", expand=True, padx=5)

        # ===== DEFINICIONES DE TOOLTIPS =====
        self.tooltips = {
            "Crear Punto de Restauración": "Crea un punto de restauración del sistema antes de hacer cambios. ✅ SEGURO.",
            "Eliminar Archivos Temporales": "Limpia archivos temporales de Windows y apps. ✅ SEGURO.",
            "Desactivar Telemetría": "Reduce el envío de datos de uso a Microsoft. ✅ SEGURO.",
            "Desactivar GameDVR": "Desactiva la grabación de juegos en segundo plano. ✅ SEGURO (mejora FPS).",
            "Desactivar Hibernación": "Libera espacio en disco (elimina archivo hiberfil.sys). ⚠️ MODERADO (no afecta laptops que usan hibernación).",
            "Desactivar Rastreo de Ubicación": "Desactiva el seguimiento de ubicación del sistema. ✅ SEGURO.",
            "Desactivar Wi-Fi Sense": "Evita que Windows comparte redes Wi-Fi automáticamente. ✅ SEGURO.",
            "Ejecutar Limpieza de Disco": "Limpia archivos innecesarios del sistema. ✅ SEGURO.",
            "Desinstalar Edge": "Intenta desinstalar Microsoft Edge (puede fallar si está protegido). ⚠️ MODERADO.",
            "Desactivar Historial de Actividad": "Evita que Windows guarde tu historial de actividades. ✅ SEGURO.",
            "Desactivar Homegroup": "Desactiva función obsoleta de redes hogareñas. ✅ SEGURO.",
            "Activar Tarea de Eliminación con clic derecho": "Agrega opción 'Eliminar' permanente en menú contextual. ✅ SEGURO.",
            "Desactivar IPv6": "Desactiva el protocolo IPv6 (puede romper conectividad en redes modernas). ❌ PELIGROSO — solo si sabes lo que haces.",
            "Preferir IPv4 sobre IPv6": "Prioriza IPv4 sin desactivar IPv6. ✅ SEGURO.",
            "Desactivar Teredo": "Desactiva túnel IPv6 sobre IPv4 (útil si no lo usas). ✅ SEGURO.",
            "Desactivar Aplicaciones en Segundo Plano": "Evita que apps como Mail o Calendario consuman recursos. ✅ SEGURO.",
            "Desactivar Optimizaciones de Pantalla Completa": "Mejora compatibilidad en juegos antiguos. ✅ SEGURO.",
            "Desactivar Copia de Microsoft (LMS)": "Desactiva servicio de telemetría empresarial. ✅ SEGURO.",
            "Desactivar Intel MM (vPro LMS)": "Desactiva servicio de administración remota de Intel. ✅ SEGURO (si no usas vPro).",
            "Eliminar OneDrive": "Desinstala OneDrive del sistema. ⚠️ MODERADO (pierdes sincronización automática).",
            "Desactivar Bandeja de Notificaciones": "Oculta iconos de sistema en la bandeja. ⚠️ MODERADO (puede ocultar actualizaciones importantes).",
            "Desactivar WPBT": "Desactiva carga de software desde firmware (mejora seguridad). ⚠️ MODERADO (solo para usuarios avanzados).",
            "Tema Oscuro para Windows": "Activa el modo oscuro en todo el sistema. ✅ SEGURO.",
            "Mostrar archivos ocultos": "Muestra archivos y carpetas ocultas en el Explorador. ✅ SEGURO.",
            "Mostrar extensiones de archivo": "Muestra la extensión de todos los archivos (.txt, .exe, etc.). ✅ SEGURO (recomendado).",
            "Centrar elementos de la barra de tareas": "Centra los iconos en la barra de tareas (estilo Windows 11). ✅ SEGURO.",
            "Botón de búsqueda en la barra de tareas": "Muestra/oculta el botón de búsqueda. ✅ SEGURO.",
            "Botón de Vista de Tareas en la barra de tareas": "Muestra/oculta el botón de escritorios virtuales. ✅ SEGURO.",
            "Botón de Widgets en la barra de tareas": "Muestra/oculta el botón de widgets. ✅ SEGURO.",
            "Ventana Adherida (Snap)": "Activa la función de ventanas adheridas. ✅ SEGURO.",
            "Diseño de Ayuda de Adherencia (Snap Assist)": "Muestra sugerencias al usar Snap. ✅ SEGURO.",
            "NumLock al inicio": "Activa NumLock automáticamente al iniciar sesión. ✅ SEGURO.",
            "Aceleración del Ratón": "Activa la aceleración del puntero (no recomendado para juegos de precisión). ⚠️ MODERADO.",
            "Teclas Pegajosas (Sticky Keys)": "Facilita usar combinaciones de teclas (Shift, Ctrl, Alt). ✅ SEGURO (para accesibilidad).",
            "plan_energia_juegos": "Configura el plan de energía en 'Alto rendimiento' para máxima potencia. ✅ SEGURO (ideal para PC enchufada).",
            "optimizar_ssd": "Verifica que TRIM esté activo (necesario para mantener el rendimiento del SSD). ✅ SEGURO.",
            "desactivar_desfrag_ssd": "Desactiva la desfragmentación automática (innecesaria y dañina para SSD). ✅ SEGURO.",
            "prioridad_gpu_cpu": "Asigna GPU de alto rendimiento y prioridad de CPU a las apps que selecciones. ✅ SEGURO (genera scripts .bat personalizados).",
            "desactivar_notificaciones": "Desactiva notificaciones y activa modo 'No molestar' durante juegos. ✅ SEGURO.",
            "script_completo": "Aplica optimizaciones comunes: desactiva Game Bar, animaciones, limpia caché, etc. ✅ SEGURO.",
        }

        # >>>>> COLUMNA IZQUIERDA <<<<<
        frame_essential = ttk.LabelFrame(left_col, text="🔧 Ajustes Esenciales", padding=10)
        frame_essential.pack(fill='x', pady=5)

        self.essential_vars = {}
        for option in [
            "Crear Punto de Restauración",
            "Eliminar Archivos Temporales",
            "Desactivar Telemetría",
            "Desactivar GameDVR",
            "Desactivar Hibernación",
            "Desactivar Rastreo de Ubicación",
            "Desactivar Wi-Fi Sense",
            "Ejecutar Limpieza de Disco",
            "Desinstalar Edge",
            "Desactivar Historial de Actividad",
            "Desactivar Homegroup",
            "Activar Tarea de Eliminación con clic derecho",
        ]:
            var = tk.BooleanVar()
            cb = ttk.Checkbutton(frame_essential, text=option, variable=var)
            cb.pack(anchor='w', padx=5, pady=2)
            self.essential_vars[option] = var
            self.create_tooltip(cb, self.tooltips.get(option, "Sin descripción."))

        frame_advanced = ttk.LabelFrame(left_col, text="⚠️ Ajustes Avanzados", padding=10)
        frame_advanced.pack(fill='x', pady=5)

        self.advanced_vars = {}
        for option in [
            "Desactivar IPv6",
            "Preferir IPv4 sobre IPv6",
            "Desactivar Teredo",
            "Desactivar Aplicaciones en Segundo Plano",
            "Desactivar Optimizaciones de Pantalla Completa",
            "Desactivar Copia de Microsoft (LMS)",
            "Desactivar Intel MM (vPro LMS)",
            "Eliminar OneDrive",
            "Desactivar Bandeja de Notificaciones",
            "Desactivar WPBT",
        ]:
            var = tk.BooleanVar()
            cb = ttk.Checkbutton(frame_advanced, text=option, variable=var)
            cb.pack(anchor='w', padx=5, pady=2)
            self.advanced_vars[option] = var
            self.create_tooltip(cb, self.tooltips.get(option, "Sin descripción."))

        # >>>>> COLUMNA DERECHA <<<<<
        frame_preferences = ttk.LabelFrame(right_col, text="🎨 Preferencias", padding=10)
        frame_preferences.pack(fill='x', pady=5)

        self.preferences_vars = {}
        for option, default in {
            "Tema Oscuro para Windows": True,
            "Mostrar archivos ocultos": True,
            "Mostrar extensiones de archivo": True,
            "Centrar elementos de la barra de tareas": True,
            "Botón de búsqueda en la barra de tareas": True,
            "Botón de Vista de Tareas en la barra de tareas": True,
            "Botón de Widgets en la barra de tareas": True,
            "Ventana Adherida (Snap)": True,
            "Diseño de Ayuda de Adherencia (Snap Assist)": True,
            "NumLock al inicio": False,
            "Aceleración del Ratón": True,
            "Teclas Pegajosas (Sticky Keys)": False,
        }.items():
            var = tk.BooleanVar(value=default)
            cb = ttk.Checkbutton(frame_preferences, text=option, variable=var)
            cb.pack(anchor='w', padx=5, pady=2)
            self.preferences_vars[option] = var
            self.create_tooltip(cb, self.tooltips.get(option, "Sin descripción."))

        # ===== OPTIMIZACIÓN PARA JUEGOS =====
        frame_gaming = ttk.LabelFrame(right_col, text="🎮 Optimización para Juegos", padding=10)
        frame_gaming.pack(fill='x', pady=5)

        self.gaming_vars = {}
        gaming_options = [
            ("⚡ Plan de Energía: Alto Rendimiento", "plan_energia_juegos"),
            ("💾 Optimizar SSD (TRIM activo)", "optimizar_ssd"),
            ("🚫 Desactivar Desfragmentación en SSD", "desactivar_desfrag_ssd"),
            ("🎮 Asignar GPU/CPU Alta a Apps", "prioridad_gpu_cpu"),
            ("🔕 Desactivar Notificaciones + Modo Enfoque", "desactivar_notificaciones"),
            ("🔧 Script Completo (GameBar, Animaciones, etc.)", "script_completo"),
        ]

        for label, key in gaming_options:
            self.gaming_vars[key] = tk.BooleanVar()
            cb = ttk.Checkbutton(frame_gaming, text=label, variable=self.gaming_vars[key])
            cb.pack(anchor='w', padx=5, pady=2)
            self.create_tooltip(cb, self.tooltips.get(key, "Sin descripción."))

        btn_select_apps = ttk.Button(frame_gaming, text="📂 Seleccionar .exe para optimizar", command=self.select_apps_for_priority)
        btn_select_apps.pack(pady=5)
        self.create_tooltip(btn_select_apps, "Selecciona los ejecutables de tus juegos o apps para asignarles prioridad ALTA de CPU y GPU de alto rendimiento.")

        # >>> NUEVO: Botón independiente para abrir guía de GPU (último elemento) <<<
        btn_gpu_guide = ttk.Button(frame_gaming, text="🖥️ Ajustes Recomendados NVIDIA/AMD/spacedesk", command=self.open_gpu_guide_window)
        btn_gpu_guide.pack(pady=(15, 5))
        self.create_tooltip(btn_gpu_guide, "Abre una guía visual paso a paso para configurar tu GPU específica. ✅ SEGURO.")

        # ===== BOTÓN APLICAR =====
        apply_frame = ttk.Frame(scrollable_frame)
        apply_frame.pack(pady=20)
        apply_btn = ttk.Button(apply_frame, text="🚀 Aplicar Cambios Seleccionados", command=self.apply_settings, width=50)
        apply_btn.pack()
        self.create_tooltip(apply_btn, "Aplica todos los cambios marcados. Algunos requieren reiniciar el PC. ⚠️ Ejecuta como Administrador.")

        self.optimizador_status = tk.StringVar(value="Listo. Selecciona los ajustes que deseas aplicar.")
        ttk.Label(scrollable_frame, textvariable=self.optimizador_status, anchor="w").pack(fill="x", padx=10, pady=5)

    def open_gpu_guide_window(self):
        """Abre ventana con guía de optimización para GPU."""
        try:
            import wmi
            c = wmi.WMI()
            gpus = c.Win32_VideoController()
            gpu_list = []
            for gpu in gpus:
                info = {
                    "Name": gpu.Name.strip() if gpu.Name else "Desconocido",
                    "Manufacturer": gpu.AdapterCompatibility.strip() if gpu.AdapterCompatibility else "Desconocido",
                    "DriverVersion": gpu.DriverVersion.strip() if gpu.DriverVersion else "Desconocido",
                    "Resolution": f"{gpu.CurrentHorizontalResolution}x{gpu.CurrentVerticalResolution}" if gpu.CurrentHorizontalResolution else "Desconocida"
                }
                gpu_list.append(info)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo obtener información de la GPU:\n{e}")
            return

        if not gpu_list:
            messagebox.showwarning("GPU no detectada", "No se encontró ninguna GPU.")
            return

        guide_win = tk.Toplevel(self)
        guide_win.title("🎮 Guía de Optimización para tu GPU")
        guide_win.geometry("900x700")
        guide_win.resizable(True, True)

        # Asegurar que esté encima
        guide_win.grab_set()
        guide_win.focus_force()
        guide_win.lift()
        guide_win.attributes('-topmost', True)
        guide_win.after_idle(guide_win.attributes, '-topmost', False)

        # Scroll
        canvas = tk.Canvas(guide_win)
        scrollbar = ttk.Scrollbar(guide_win, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Título
        ttk.Label(scrollable_frame, text="📊 Información y Guía de Optimización", font=("Segoe UI", 16, "bold")).pack(pady=20)

        # Mostrar cada GPU con su guía personalizada
        for i, gpu in enumerate(gpu_list):
            frame = ttk.LabelFrame(scrollable_frame, text=f"GPU #{i+1}: {gpu['Name']}", padding=10)
            frame.pack(fill='x', padx=20, pady=10)

            info_text = f"""
Fabricante: {gpu['Manufacturer']}
Driver: {gpu['DriverVersion']}
Resolución: {gpu['Resolution']}
"""
            ttk.Label(frame, text=info_text, font=("Consolas", 10), justify="left").pack(anchor="w", pady=5)

            # Detectar si es spacedesk
            is_spacedesk = "spacedesk" in gpu['Name'].lower() or "spacedesk" in gpu['Manufacturer'].lower()
            is_nvidia = "NVIDIA" in gpu['Name'].upper() or "NVIDIA" in gpu['Manufacturer'].upper()
            is_amd = "AMD" in gpu['Name'].upper() or "AMD" in gpu['Manufacturer'].upper() or "RADEON" in gpu['Name'].upper()

            if is_spacedesk:
                self.add_spacedesk_guide(frame)
            elif is_nvidia:
                self.add_nvidia_guide(frame)
            elif is_amd:
                self.add_amd_guide(frame)
            else:
                ttk.Label(frame, text="ℹ️ Guía no disponible para este modelo.", font=("Segoe UI", 9, "italic")).pack(pady=5)

        ttk.Button(scrollable_frame, text="✅ Entendido - Cerrar", command=guide_win.destroy, style="Accent.TButton").pack(pady=30)

    def add_spacedesk_guide(self, parent_frame):
        ttk.Label(parent_frame, text="🔧 Guía de Optimización para spacedesk", font=("Segoe UI", 11, "bold")).pack(pady=(15,5), anchor="w")
        steps = [
            ("spacedesk es un software que convierte dispositivos en pantallas secundarias.", None),
            ("No requiere optimización de rendimiento como una GPU de juego.", None),
            ("Si experimentas problemas, actualiza los drivers o revisa la documentación.", None),
        ]
        for i, (text, url) in enumerate(steps, start=1):
            step_frame = ttk.Frame(parent_frame)
            step_frame.pack(fill='x', pady=4)
            num_label = ttk.Label(step_frame, text=f"{i}.", font=("Segoe UI", 9, "bold"), width=3)
            num_label.pack(side="left", anchor="nw", padx=(0, 8))
            desc_label = ttk.Label(step_frame, text=text, font=("Segoe UI", 9), wraplength=650, justify="left")
            desc_label.pack(side="left", anchor="nw")
            if url:
                btn = ttk.Button(step_frame, text="👁️ Ver Guía", command=lambda u=url: webbrowser.open(u))
                btn.pack(side="right", padx=(10, 0))

        ttk.Label(parent_frame, text="📚 Recursos Oficiales:", font=("Segoe UI", 10, "bold")).pack(pady=(15,5), anchor="w")
        resources = [
            ("Documentación Oficial", "https://www.spacedesk.net/documentation/"),
            ("Descargar Drivers", "https://www.spacedesk.net/download/"),
        ]
        for title, link in resources:
            res_frame = ttk.Frame(parent_frame)
            res_frame.pack(fill='x', pady=3)
            ttk.Label(res_frame, text=title, font=("Segoe UI", 9)).pack(side="left", anchor="w")
            ttk.Button(res_frame, text="🔗 Abrir", command=lambda l=link: webbrowser.open(l)).pack(side="right", padx=(10,0))

    def add_nvidia_guide(self, parent_frame):
        ttk.Label(parent_frame, text="🔧 Guía de Optimización para NVIDIA", font=("Segoe UI", 11, "bold")).pack(pady=(15,5), anchor="w")
        steps = [
            ("Abre el Panel de Control de NVIDIA", "https://www.nvidia.com/es-es/geforce/forums/game-ready-drivers/13/256597/how-to-open-nvidia-control-panel/"),
            ("Ve a 'Administrar configuración 3D' > 'Ajustes de programa'", None),
            ("Selecciona tu programa o agrega el .exe", None),
            ("'Modo de energía preferido' → 'Máximo rendimiento'", None),
            ("'Latencia de baja latencia' → 'Ultra'", None),
            ("'Shader Cache' → 'On'", None),
            ("'Procesador gráfico preferido' → 'GPU de alto rendimiento NVIDIA'", None),
            ("Haz clic en 'Aplicar'", None),
        ]
        for i, (text, url) in enumerate(steps, start=1):
            step_frame = ttk.Frame(parent_frame)
            step_frame.pack(fill='x', pady=4)
            num_label = ttk.Label(step_frame, text=f"{i}.", font=("Segoe UI", 9, "bold"), width=3)
            num_label.pack(side="left", anchor="nw", padx=(0, 8))
            desc_label = ttk.Label(step_frame, text=text, font=("Segoe UI", 9), wraplength=650, justify="left")
            desc_label.pack(side="left", anchor="nw")
            if url:
                btn = ttk.Button(step_frame, text="👁️ Ver Guía", command=lambda u=url: webbrowser.open(u))
                btn.pack(side="right", padx=(10, 0))

        ttk.Label(parent_frame, text="📚 Recursos Oficiales:", font=("Segoe UI", 10, "bold")).pack(pady=(15,5), anchor="w")
        resources = [
            ("Configuración 3D - NVIDIA", "https://nvidia.custhelp.com/app/answers/detail/a_id/5258"),
            ("Tecnología de Baja Latencia", "https://www.nvidia.com/es-es/geforce/technologies/low-latency-mode/"),
            ("Shader Cache Explicado", "https://www.nvidia.com/es-es/geforce/technologies/shader-cache/"),
        ]
        for title, link in resources:
            res_frame = ttk.Frame(parent_frame)
            res_frame.pack(fill='x', pady=3)
            ttk.Label(res_frame, text=title, font=("Segoe UI", 9)).pack(side="left", anchor="w")
            ttk.Button(res_frame, text="🔗 Abrir", command=lambda l=link: webbrowser.open(l)).pack(side="right", padx=(10,0))

    def add_amd_guide(self, parent_frame):
        ttk.Label(parent_frame, text="🔧 Guía de Optimización para AMD", font=("Segoe UI", 11, "bold")).pack(pady=(15,5), anchor="w")
        steps = [
            ("Abre AMD Radeon Software (Win + R → 'RadeonSoftware')", "https://www.amd.com/es/support/kb/faq/gs-101"),
            ("Ve a la pestaña 'Juego'", None),
            ("Selecciona tu juego o agrega el .exe", None),
            ("'Modo gráfico' → 'Optimizado para rendimiento'", None),
            ("Activa 'Anti-Lag'", "https://www.amd.com/es/products/software/adrenalin/radeon-software-anti-lag.html"),
            ("Activa 'Radeon Boost' (si tu juego lo soporta)", "https://www.amd.com/es/products/software/adrenalin/radeon-boost.html"),
            ("'Espera de fotogramas' → 'Desactivado'", None),
            ("Haz clic en 'Aplicar'", None),
        ]
        for i, (text, url) in enumerate(steps, start=1):
            step_frame = ttk.Frame(parent_frame)
            step_frame.pack(fill='x', pady=4)
            num_label = ttk.Label(step_frame, text=f"{i}.", font=("Segoe UI", 9, "bold"), width=3)
            num_label.pack(side="left", anchor="nw", padx=(0, 8))
            desc_label = ttk.Label(step_frame, text=text, font=("Segoe UI", 9), wraplength=650, justify="left")
            desc_label.pack(side="left", anchor="nw")
            if url:
                btn = ttk.Button(step_frame, text="👁️ Ver Guía", command=lambda u=url: webbrowser.open(u))
                btn.pack(side="right", padx=(10, 0))

        ttk.Label(parent_frame, text="📚 Recursos Oficiales:", font=("Segoe UI", 10, "bold")).pack(pady=(15,5), anchor="w")
        resources = [
            ("AMD Radeon Anti-Lag", "https://www.amd.com/es/products/software/adrenalin/radeon-software-anti-lag.html"),
            ("AMD Radeon Boost", "https://www.amd.com/es/products/software/adrenalin/radeon-boost.html"),
            ("Guía de Configuración 3D - AMD", "https://www.amd.com/es/support/kb/faq/dh-012"),
        ]
        for title, link in resources:
            res_frame = ttk.Frame(parent_frame)
            res_frame.pack(fill='x', pady=3)
            ttk.Label(res_frame, text=title, font=("Segoe UI", 9)).pack(side="left", anchor="w")
            ttk.Button(res_frame, text="🔗 Abrir", command=lambda l=link: webbrowser.open(l)).pack(side="right", padx=(10,0))

    def select_apps_for_priority(self):
        filepaths = filedialog.askopenfilenames(
            title="Selecciona los .exe de tus juegos o apps",
            filetypes=[("Ejecutables", "*.exe")]
        )
        if filepaths:
            self.selected_apps = list(filepaths)
            messagebox.showinfo("Apps Seleccionadas", f"Seleccionaste {len(self.selected_apps)} apps:\n" + "\n".join([os.path.basename(f) for f in self.selected_apps]))

    def apply_settings(self):
        if not self.has_admin_privileges():
            messagebox.showerror("Error", "⚠️ Este script requiere ejecutarse como Administrador.")
            return

        if not messagebox.askyesno("Confirmar", "¿Estás seguro de aplicar estos cambios? Algunos requieren reinicio."):
            return

        def worker():
            try:
                self.apply_essential_tweaks()
                self.apply_advanced_tweaks()
                self.apply_preferences_tweaks()
                self.apply_gaming_tweaks()
                self.after(0, lambda: messagebox.showinfo("✅ Éxito", "¡Cambios aplicados! Reinicia tu PC para ver todos los efectos."))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("❌ Error", f"Error al aplicar cambios:\n{str(e)}"))

        threading.Thread(target=worker, daemon=True).start()

    def has_admin_privileges(self):
        try:
            return os.getuid() == 0
        except AttributeError:
            import ctypes
            return ctypes.windll.shell32.IsUserAnAdmin() != 0

    def apply_essential_tweaks(self):
        if self.essential_vars.get("Desactivar Telemetría", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKLM\SOFTWARE\Policies\Microsoft\Windows\DataCollection", "/v", "AllowTelemetry", "/t", "REG_DWORD", "/d", "0", "/f"], shell=True)
        if self.essential_vars.get("Desactivar GameDVR", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\GameDVR", "/v", "AppCaptureEnabled", "/t", "REG_DWORD", "/d", "0", "/f"], shell=True)
        if self.essential_vars.get("Eliminar Archivos Temporales", tk.BooleanVar()).get():
            subprocess.run(["cleanmgr", "/sagerun:1"], shell=True)
        if self.essential_vars.get("Desactivar Hibernación", tk.BooleanVar()).get():
            subprocess.run(["powercfg", "/h", "off"], shell=True)
        if self.essential_vars.get("Desinstalar Edge", tk.BooleanVar()).get():
            # ✅ Solución: Añadir --silent para evitar prompts interactivos
            subprocess.run(["winget", "uninstall", "Microsoft.Edge", "--silent"], shell=True, capture_output=True)

    def apply_advanced_tweaks(self):
        if self.advanced_vars.get("Desactivar IPv6", tk.BooleanVar()).get():
            subprocess.run(["netsh", "interface", "ipv6", "set", "state", "disabled"], shell=True)
        if self.advanced_vars.get("Preferir IPv4 sobre IPv6", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKLM\SYSTEM\CurrentControlSet\Services\Tcpip6\Parameters", "/v", "DisabledComponents", "/t", "REG_DWORD", "/d", "32", "/f"], shell=True)
        if self.advanced_vars.get("Desactivar Teredo", tk.BooleanVar()).get():
            subprocess.run(["netsh", "interface", "teredo", "set", "state", "disabled"], shell=True)
        if self.advanced_vars.get("Desactivar Aplicaciones en Segundo Plano", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\BackgroundApps", "/v", "GlobalUserDisabled", "/t", "REG_DWORD", "/d", "1", "/f"], shell=True)
        if self.advanced_vars.get("Desactivar Optimizaciones de Pantalla Completa", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\GameDVR", "/v", "AllowFullScreenOptimization", "/t", "REG_DWORD", "/d", "0", "/f"], shell=True)
        if self.advanced_vars.get("Desactivar Copia de Microsoft (LMS)", tk.BooleanVar()).get():
            subprocess.run(["sc", "stop", "LMS"], shell=True)
            subprocess.run(["sc", "config", "LMS", "start=", "disabled"], shell=True)
        if self.advanced_vars.get("Desactivar Intel MM (vPro LMS)", tk.BooleanVar()).get():
            subprocess.run(["sc", "stop", "Intel(R) Management and Security Application"], shell=True)
            subprocess.run(["sc", "config", "Intel(R) Management and Security Application", "start=", "disabled"], shell=True)
        if self.advanced_vars.get("Eliminar OneDrive", tk.BooleanVar()).get():
            subprocess.run(["taskkill", "/f", "/im", "OneDrive.exe"], shell=True)
            subprocess.run([r"%SystemRoot%\SysWOW64\OneDriveSetup.exe", "/uninstall"], shell=True)
        if self.advanced_vars.get("Desactivar Bandeja de Notificaciones", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\PushNotifications", "/v", "ToastEnabled", "/t", "REG_DWORD", "/d", "0", "/f"], shell=True)
        if self.advanced_vars.get("Desactivar WPBT", tk.BooleanVar()).get():
            subprocess.run(["bcdedit", "/set", "wpbt", "disabled"], shell=True)

    def apply_preferences_tweaks(self):
        if self.preferences_vars.get("Tema Oscuro para Windows", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\Themes\Personalize", "/v", "AppsUseLightTheme", "/t", "REG_DWORD", "/d", "0", "/f"], shell=True)
        if self.preferences_vars.get("Mostrar archivos ocultos", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\Explorer\Advanced", "/v", "Hidden", "/t", "REG_DWORD", "/d", "1", "/f"], shell=True)
        if self.preferences_vars.get("Mostrar extensiones de archivo", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\Explorer\Advanced", "/v", "HideFileExt", "/t", "REG_DWORD", "/d", "0", "/f"], shell=True)
        if self.preferences_vars.get("Centrar elementos de la barra de tareas", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\Explorer\Advanced", "/v", "TaskbarAl", "/t", "REG_DWORD", "/d", "1", "/f"], shell=True)
        if self.preferences_vars.get("Botón de búsqueda en la barra de tareas", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\Search", "/v", "SearchboxTaskbarMode", "/t", "REG_DWORD", "/d", "1", "/f"], shell=True)
        if self.preferences_vars.get("Botón de Vista de Tareas en la barra de tareas", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\Explorer\Advanced", "/v", "ShowTaskViewButton", "/t", "REG_DWORD", "/d", "1", "/f"], shell=True)
        if self.preferences_vars.get("Botón de Widgets en la barra de tareas", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\Explorer\Advanced", "/v", "TaskbarDa", "/t", "REG_DWORD", "/d", "1", "/f"], shell=True)
        if self.preferences_vars.get("Ventana Adherida (Snap)", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\Explorer\Advanced", "/v", "EnableSnapAssistFlyout", "/t", "REG_DWORD", "/d", "1", "/f"], shell=True)
        if self.preferences_vars.get("Diseño de Ayuda de Adherencia (Snap Assist)", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\Explorer\Advanced", "/v", "SnapAssist", "/t", "REG_DWORD", "/d", "1", "/f"], shell=True)
        if self.preferences_vars.get("NumLock al inicio", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Control Panel\Keyboard", "/v", "InitialKeyboardIndicators", "/t", "REG_SZ", "/d", "2", "/f"], shell=True)
        if self.preferences_vars.get("Aceleración del Ratón", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Control Panel\Mouse", "/v", "MouseSpeed", "/t", "REG_SZ", "/d", "1", "/f"], shell=True)
            subprocess.run(["reg", "add", r"HKCU\Control Panel\Mouse", "/v", "MouseThreshold1", "/t", "REG_SZ", "/d", "6", "/f"], shell=True)
            subprocess.run(["reg", "add", r"HKCU\Control Panel\Mouse", "/v", "MouseThreshold2", "/t", "REG_SZ", "/d", "10", "/f"], shell=True)
        if self.preferences_vars.get("Teclas Pegajosas (Sticky Keys)", tk.BooleanVar()).get():
            subprocess.run(["reg", "add", r"HKCU\Control Panel\Accessibility\StickyKeys", "/v", "Flags", "/t", "REG_SZ", "/d", "506", "/f"], shell=True)

    def apply_gaming_tweaks(self):
        if self.gaming_vars["plan_energia_juegos"].get():
            subprocess.run(["powercfg", "-setactive", "8c5e7fda-e8bf-4a96-9a85-a6e23a8c635c"], shell=True)
        if self.gaming_vars["optimizar_ssd"].get():
            result = subprocess.run(["fsutil", "behavior", "query", "DisableDeleteNotify"], capture_output=True, text=True, shell=True)
            if "DisableDeleteNotify = 0" not in result.stdout:
                subprocess.run(["fsutil", "behavior", "set", "DisableDeleteNotify", "0"], shell=True)
        if self.gaming_vars["desactivar_desfrag_ssd"].get():
            subprocess.run(["schtasks", "/change", "/tn", r"\Microsoft\Windows\Defrag\ScheduledDefrag", "/disable"], shell=True)
        if self.gaming_vars["prioridad_gpu_cpu"].get():
            self.generate_priority_script()
        if self.gaming_vars["desactivar_notificaciones"].get():
            subprocess.run(["reg", "add", r"HKCU\Software\Microsoft\Windows\CurrentVersion\PushNotifications", "/v", "ToastEnabled", "/t", "REG_DWORD", "/d", "0", "/f"], shell=True)
        if self.gaming_vars["script_completo"].get():
            self.generate_complete_script()

    def generate_priority_script(self):
        """Genera un script .bat para asignar prioridad alta a apps seleccionadas."""
        if not hasattr(self, 'selected_apps') or not self.selected_apps:
            messagebox.showwarning("Advertencia", "Primero selecciona apps para optimizar.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        script_name = f"priority-high-{timestamp}.bat"
        dest_folder = filedialog.askdirectory(title="Selecciona carpeta destino para el script")
        if not dest_folder:
            return

        script_path = os.path.join(dest_folder, script_name)

        try:
            with open(script_path, "w", encoding="utf-8") as f:
                f.write("@echo off\n")
                f.write(f"REM Script generado por CSM v.2.0 el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("REM Asigna prioridad ALTA de CPU y GPU a las apps seleccionadas.\n")
                f.write("echo Asignando prioridad alta...\n\n")
                for path in self.selected_apps:
                    f.write(f'wmic process where name="{os.path.basename(path)}" CALL setpriority 128\n')
                f.write('\necho ¡Prioridad asignada! ¡Listo para jugar!\npause\n')
            messagebox.showinfo("Script generado", f"Script guardado en:\n{script_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def generate_complete_script(self):
        """Genera un script completo de optimización."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        script_name = f"optimizacion-completa-{timestamp}.bat"
        dest_folder = filedialog.askdirectory(title="Selecciona carpeta destino para el script")
        if not dest_folder:
            return

        script_path = os.path.join(dest_folder, script_name)

        try:
            with open(script_path, "w", encoding="utf-8") as f:
                f.write("@echo off\n")
                f.write(f"REM Script de optimización completo generado por CSM v.2.0 el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("REM Aplica ajustes esenciales y avanzados para juegos.\n")
                f.write("echo Aplicando optimizaciones...\n\n")

                # Desactivar Game Bar
                f.write("reg add HKCU\\Software\\Microsoft\\GameBar /v ShowStartupPanel /t REG_DWORD /d 0 /f\n")
                f.write("reg add HKCU\\Software\\Microsoft\\GameBar /v GamePanelStartupTipIndex /t REG_DWORD /d 3 /f\n")
                f.write("reg add HKCU\\Software\\Microsoft\\GameBar /v AllowAutoGameMode /t REG_DWORD /d 0 /f\n")
                f.write("reg add HKCU\\Software\\Microsoft\\GameBar /v UseNexus /t REG_DWORD /d 0 /f\n")

                # Desactivar animaciones
                f.write("reg add HKCU\\Control Panel\\Desktop\\WindowMetrics /v MinAnimate /t REG_SZ /d 0 /f\n")
                f.write("reg add HKCU\\Software\\Microsoft\\Windows\\CurrentVersion\\Explorer\\Advanced /v TaskbarAnimations /t REG_DWORD /d 0 /f\n")
                f.write("reg add HKCU\\Software\\Microsoft\\Windows\\CurrentVersion\\Explorer\\Advanced /v ListviewAlphaSelect /t REG_DWORD /d 0 /f\n")
                f.write("reg add HKCU\\Software\\Microsoft\\Windows\\CurrentVersion\\Explorer\\Advanced /v ListviewShadow /t REG_DWORD /d 0 /f\n")

                # Limpiar caché
                f.write("echo Limpiando caché...\n")
                f.write("del /q /f /s %temp%\\*\n")
                f.write("del /q /f /s %windir%\\Temp\\*\n")
                f.write("echo Caché limpiado.\n\n")

                f.write("echo ¡Optimizaciones aplicadas! ¡Listo para jugar!\npause\n")
            messagebox.showinfo("Script generado", f"Script guardado en:\n{script_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ---------------- Pestaña "Repo-Software" ----------------
    def _build_repo_tab(self):
        frame = self.tab_repo

        ttk.Label(frame, text="Backup & Restauración con Winget", font=("Segoe UI", 14, "bold")).pack(pady=10)

        btn_frame = ttk.Frame(frame); btn_frame.pack(pady=10)

        ttk.Button(btn_frame, text="Listar Software Instalado", command=self.list_installed_software, width=25).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Generar Backup (.bat + .json + .xlsx + .zip)", command=self.generate_full_backup, width=35).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Abrir winstall.app", command=self.open_winstall, width=20).pack(side="left", padx=5)

        ttk.Label(frame, text="Software detectado (winget list):").pack(anchor="w", padx=10, pady=(10, 0))
        text_frame = ttk.Frame(frame)
        text_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.software_text = tk.Text(text_frame, wrap="none", bg="#2a2a2a", fg="white", font=("Consolas", 10))
        self.software_text.pack(side="left", fill="both", expand=True)

        v_scroll = ttk.Scrollbar(text_frame, orient="vertical", command=self.software_text.yview)
        v_scroll.pack(side="right", fill="y")
        h_scroll = ttk.Scrollbar(frame, orient="horizontal", command=self.software_text.xview)
        h_scroll.pack(fill="x", padx=10)
        self.software_text.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.repo_status = tk.StringVar(value="Listo. Usa 'Listar Software Instalado' para comenzar.")
        ttk.Label(frame, textvariable=self.repo_status, anchor="w").pack(fill="x", padx=10, pady=5)

    # ---------------- Pestaña "Pckgr / Intune" ----------------
    def _build_pckgr_tab(self):
        frame = self.tab_pckgr
        frame.pack(fill="both", expand=True, padx=20, pady=20)

        ttk.Label(frame, text="Integración con Pckgr para Intune", font=("Segoe UI", 16, "bold")).pack(anchor="nw", pady=(0, 10))
        quote_label = ttk.Label(frame, text=quote, wraplength=800, justify="center", font=("Segoe UI", 11, "italic"))
        quote_label.pack(pady=10)

        steps = ttk.LabelFrame(frame, text="Genera tu backup para Pckgr", padding=15)
        steps.pack(fill="x", pady=10)

        instructions = (
            "1. Haz clic en 'Listar Software Instalado' para ver tus apps.\n"
            "2. Luego, haz clic en 'Generar Backup (.bat + .json + .xlsx + .zip)'.\n"
            "3. Selecciona una carpeta de destino.\n"
            "4. Sube el archivo .json generado a Pckgr en winstall.app.\n"
            "5. ¡Listo! Despliega en Intune con actualizaciones automáticas."
        )
        ttk.Label(steps, text=instructions, justify="left", font=("Segoe UI", 10)).pack(anchor="nw", pady=(6,2))

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=10)

        ttk.Button(btn_frame, text="Listar Software Instalado", command=self.list_installed_software, width=25).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Generar Backup Completo", command=self.generate_full_backup, width=30).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Abrir winstall.app", command=self.open_winstall, width=20).pack(side="left", padx=5)

        ttk.Label(frame, text="Software detectado (winget list):").pack(anchor="w", padx=10, pady=(10, 0))
        text_frame = ttk.Frame(frame)
        text_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.software_text = tk.Text(text_frame, wrap="none", bg="#2a2a2a", fg="white", font=("Consolas", 10))
        self.software_text.pack(side="left", fill="both", expand=True)

        v_scroll = ttk.Scrollbar(text_frame, orient="vertical", command=self.software_text.yview)
        v_scroll.pack(side="right", fill="y")
        h_scroll = ttk.Scrollbar(frame, orient="horizontal", command=self.software_text.xview)
        h_scroll.pack(fill="x", padx=10)
        self.software_text.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.repo_status = tk.StringVar(value="Listo. Usa 'Listar Software Instalado' para comenzar.")
        ttk.Label(frame, textvariable=self.repo_status, anchor="w").pack(fill="x", padx=10, pady=5)

    # ---------------- Pestaña "Temas" ----------------
    def _build_themes_tab(self):
        frame = self.tab_temas
        ttk.Label(frame, text="Selecciona un tema:", font=("Segoe UI", 12, "bold")).pack(pady=20)

        themes = ["darkly", "superhero", "cyborg", "vapor", "solar", "cosmo", "flatly", "litera", "minty", "lumen"]
        for theme in themes:
            btn = ttk.Button(frame, text=theme.capitalize(), command=lambda t=theme: self.change_theme(t), width=20)
            btn.pack(pady=5)

    # ---------------- Pestaña "Acerca de" ----------------
    def _build_about_tab(self):
        about_frame = ttk.Frame(self.tab_about); about_frame.pack(fill="both", expand=True, padx=20, pady=20)
        title_label = ttk.Label(about_frame, text="CSM – Check Software Mínimo", font=("Segoe UI", 16, "bold"))
        title_label.pack(anchor="nw", pady=(6,2))

        info = (
            "- Creado por LSCF, para uso mejorar las instalaciones semi-automaticas.\n"
            "- Ayudado por IA, idea original de LSCF.\n"
            "- Esta desarrollado para uso libre, el software que se use, key, portables son responsabilidad del usuario final no del creador.\n\n"
            "Cambios v.2.0:\n"
            " • Nueva pestaña 'Pckgr / Intune' con backup automático (Excel, JSON, ZIP).\n"
            " • Barra de progreso gráfica en 'Actualizaciones'.\n"
            " • Integración directa con Pckgr para despliegues en Intune.\n"
        )
        ttk.Label(about_frame, text=info, justify="left").pack(anchor="nw", pady=(6,2))

    # ---------------- Tooltip System ----------------
    def create_tooltip(self, widget, text):
        tooltip = None
        def show_tip(event):
            nonlocal tooltip
            if tooltip: return
            x = event.widget.winfo_rootx() + 20
            y = event.widget.winfo_rooty() + 30
            tooltip = tk.Toplevel(self)
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{x}+{y}")
            label = tk.Label(tooltip, text=text, background="lightyellow", relief="solid", borderwidth=1, padx=5, pady=3)
            label.pack()
            widget.tooltip_window = tooltip
        def hide_tip(event):
            nonlocal tooltip
            if hasattr(event.widget, 'tooltip_window'):
                event.widget.tooltip_window.destroy()
                delattr(event.widget, 'tooltip_window')
            tooltip = None
        widget.bind("<Enter>", show_tip)
        widget.bind("<Leave>", hide_tip)

    def on_tree_motion(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid or iid not in self.path_map:
            return
        full_path = self.path_map[iid]
        filename = os.path.basename(full_path).lower()
        tooltip_text = "Información no disponible."
        for key, info in SOFTWARE_INFO.items():
            if key in filename:
                tooltip_text = info
                break
        self.create_tooltip(event.widget, tooltip_text)

    # ---------------- Theme System ----------------
    def change_theme(self, theme_name):
        try:
            self.style.theme_use(theme_name)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cambiar al tema {theme_name}: {e}")

    # ---------------- Folder Selection ----------------
    def select_folder(self):
        folder = filedialog.askdirectory(initialdir=self.folder_var.get())
        if folder:
            self.folder_var.set(folder)

    # ---------------- Scanning Tree ----------------
    def scan_now(self):
        folder = self.folder_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Carpeta inválida", "Selecciona una carpeta existente.")
            return

        self.status.set("Escaneando…")
        self.update_idletasks()
        self.path_map.clear()
        self.path_type.clear()
        self.check_vars.clear()
        for iid in list(self.tree.get_children()):
            self.tree.delete(iid)
        self._insert_tree("", folder)
        self.status.set("Escaneo completo.")

    def _insert_tree(self, parent, path):
        try:
            entries = sorted(os.listdir(path), key=lambda s: s.lower())
        except Exception:
            return

        cmd_files = [e for e in entries if os.path.isfile(os.path.join(path,e)) and os.path.splitext(e)[1].lower() in ('.bat', '.cmd')]
        files_to_show = []
        if cmd_files:
            files_to_show = sorted(cmd_files, key=lambda s: s.lower())
        else:
            for e in entries:
                full = os.path.join(path,e)
                if os.path.isfile(full):
                    low = e.lower(); ext = os.path.splitext(e)[1].lower()
                    if ext in ('.zip','.rar','.txt','.key','.reg','.cmd','.bat'):
                        files_to_show.append(e)
                    elif ext in ('.exe','.msi','.msixbundle'):
                        if low.startswith(SPECIAL_PORTABLE_PREFIX) or is_minios_folder(path) or 'portable' in low:
                            files_to_show.append(e)
                        elif any(k in low for k in KEYWORDS):
                            files_to_show.append(e)

        for entry in files_to_show:
            full = os.path.join(path, entry); low = entry.lower(); ext = os.path.splitext(entry)[1].lower()
            ptype = None
            if ext in ('.cmd','.bat'): ptype = "portable"
            elif ext == '.txt': ptype = "txt"
            elif ext == '.key' or 'patch' in low: ptype = "key"
            elif ext in ('.zip','.rar'): ptype = "archive"
            elif ext == '.reg': ptype = "reg"
            elif ext in ('.exe','.msi','.msixbundle'):
                if low.startswith(SPECIAL_PORTABLE_PREFIX) or is_minios_folder(path) or 'portable' in low: ptype = "portable"
                elif any(k in low for k in KEYWORDS): ptype = "installer"
            if ptype:
                display = entry
                if 'sdio_auto' in low and ext == '.bat': display = "SDIO_ejecutable"
                iid = self.tree.insert(parent, "end", text=f"☐ {display}")
                self.path_map[iid] = full; self.path_type[iid] = ptype; self.check_vars[iid] = False

        for entry in entries:
            full = os.path.join(path, entry)
            if os.path.isdir(full) and entry.lower() != "Software Instaladores":
                node = self.tree.insert(parent, "end", text=entry, open=False)
                self._insert_tree(node, full)

    # ---------------- Tree Interactions ----------------
    def on_tree_click(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid or iid not in self.check_vars: return
        bbox = self.tree.bbox(iid, "#0")
        if bbox:
            x_offset = event.x - bbox[0]
            if x_offset < 20:
                self.toggle_check(iid)

    def toggle_check(self, iid):
        self.check_vars[iid] = not self.check_vars[iid]
        text = self.tree.item(iid, "text")
        if "SDIO_ejecutable" in text:
            symbol = "☑" if self.check_vars[iid] else "☐"; self.tree.item(iid, text=f"{symbol} SDIO_ejecutable"); return
        base = os.path.basename(self.path_map[iid]); symbol = "☑" if self.check_vars[iid] else "☐"
        self.tree.item(iid, text=f"{symbol} {base}")

    def set_all_checks(self, value: bool):
        for iid in list(self.check_vars.keys()):
            self.check_vars[iid] = value
            text = self.tree.item(iid, "text")
            if "SDIO_ejecutable" in text:
                symbol = "☑" if value else "☐"; self.tree.item(iid, text=f"{symbol} SDIO_ejecutable")
            else:
                base = os.path.basename(self.path_map[iid]); symbol = "☑" if value else "☐"; self.tree.item(iid, text=f"{symbol} {base}")

    # ---------------- Install / Run ----------------
    def install_selected(self):
        selected = [(iid,self.path_map[iid],self.path_type[iid]) for iid in self.check_vars if self.check_vars[iid] and self.path_type[iid]=="installer"]
        if not selected:
            messagebox.showinfo("Instalar", "No hay instaladores seleccionados.")
            return

        self.install_btn.config(state="disabled")
        errors = []
        self.total_tasks = len(selected)
        self.current_task = 0

        progress = ttk.Progressbar(self.tab_installers, mode="determinate", length=300, maximum=100)
        progress.pack(pady=10)
        progress_label = ttk.Label(self.tab_installers, text="0%")
        progress_label.pack()

        self.current_progress = (progress, progress_label)

        def worker():
            for i, (iid,path,ptype) in enumerate(selected, 1):
                self.current_task = i
                pct = int((i / self.total_tasks) * 100)
                self.status.set(f"Instalando {i}/{self.total_tasks}: {os.path.basename(path)}")
                if self.current_progress:
                    self.after(0, lambda p=pct: self.current_progress[0].config(value=p))
                    self.after(0, lambda t=f"{pct}%": self.current_progress[1].config(text=t))
                self.update_idletasks()
                try:
                    proc = subprocess.Popen([path])
                    proc.wait()
                except Exception as e:
                    errors.append(f"Instalar {path}: {e}")
            self.status.set("Instalaciones completadas." if not errors else "Hubo errores durante la instalación.")
            if self.current_progress:
                self.after(0, lambda: self.current_progress[0].pack_forget())
                self.after(0, lambda: self.current_progress[1].pack_forget())
                self.current_progress = None
            self.install_btn.config(state="normal")
            if errors:
                messagebox.showerror("Errores", "\n".join(errors))
            else:
                messagebox.showinfo("Instalar", "Instalaciones completadas.")

        threading.Thread(target=worker, daemon=True).start()

    def run_selected(self):
        selected = [(iid,self.path_map[iid],self.path_type[iid]) for iid in self.check_vars if self.check_vars[iid]]
        if not selected:
            messagebox.showinfo("Ejecutar", "No hay elementos seleccionados.")
            return

        self.run_btn.config(state="disabled")
        errors = []
        self.total_tasks = len(selected)
        self.current_task = 0

        progress = ttk.Progressbar(self.tab_installers, mode="determinate", length=300, maximum=100)
        progress.pack(pady=10)
        progress_label = ttk.Label(self.tab_installers, text="0%")
        progress_label.pack()

        self.current_progress = (progress, progress_label)

        def worker():
            for i, (iid,path,ptype) in enumerate(selected, 1):
                self.current_task = i
                pct = int((i / self.total_tasks) * 100)
                self.status.set(f"Ejecutando {i}/{self.total_tasks}: {os.path.basename(path)}")
                if self.current_progress:
                    self.after(0, lambda p=pct: self.current_progress[0].config(value=p))
                    self.after(0, lambda t=f"{pct}%": self.current_progress[1].config(text=t))
                self.update_idletasks()
                try:
                    if ptype == "txt": subprocess.Popen(["notepad", path])
                    elif ptype == "archive": self.handle_archive(path)
                    elif ptype == "key": self.handle_key(path)
                    else: subprocess.Popen([path], shell=True)
                except Exception as e:
                    errors.append(f"Ejecutar {path}: {e}")
            if self.current_progress:
                self.after(0, lambda: self.current_progress[0].pack_forget())
                self.after(0, lambda: self.current_progress[1].pack_forget())
                self.current_progress = None
            self.run_btn.config(state="normal")
            if errors:
                messagebox.showerror("Errores", "\n".join(errors))

        threading.Thread(target=worker, daemon=True).start()

    # ---------------- Key handling ----------------
    def handle_key(self, keypath):
        base = os.path.splitext(os.path.basename(keypath))[0]
        candidates = []
        for root in [r"C:\Program Files", r"C:\Program Files (x86)", r"C:\ProgramData", r"C:\Users"]:
            if os.path.isdir(root):
                try:
                    for d in os.listdir(root):
                        if base.lower() in d.lower():
                            candidates.append(os.path.join(root,d))
                except Exception:
                    pass
            if len(candidates) > 60:
                break
        if not candidates:
            dest = filedialog.askdirectory(title=f"Selecciona carpeta destino para {os.path.basename(keypath)}")
            if not dest: return
            try:
                shutil.copy2(keypath, os.path.join(dest, os.path.basename(keypath)))
                messagebox.showinfo("Copiado", f"Copiado a {dest}")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            return
        choice = self.ask_choice("Selecciona carpeta destino", f"Se encontraron carpetas relacionadas con '{base}'", candidates)
        if not choice: return
        if not messagebox.askyesno("Confirmar", f"Copiar {os.path.basename(keypath)} a:\n{choice} ?"): return
        try:
            shutil.copy2(keypath, os.path.join(choice, os.path.basename(keypath)))
            messagebox.showinfo("Copiado", "Archivo copiado correctamente.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def ask_choice(self, title, prompt, options):
        win = tk.Toplevel(self); win.title(title); win.geometry("720x320")
        ttk.Label(win, text=prompt).pack(pady=6)
        lb = tk.Listbox(win); lb.pack(fill="both", expand=True, padx=8, pady=8)
        for o in options: lb.insert("end", o)
        sel = {'choice': None}
        def ok():
            s = lb.curselection()
            if s: sel['choice'] = lb.get(s[0])
            win.destroy()
        def cancel():
            win.destroy()
        btns = ttk.Frame(win); btns.pack(pady=6)
        ttk.Button(btns, text="Aceptar", command=ok).pack(side="left", padx=6)
        ttk.Button(btns, text="Cancelar", command=cancel).pack(side="left", padx=6)
        self.wait_window(win); return sel['choice']

    # ---------------- Archive handling ----------------
    def handle_archive(self, archive_path):
        dest = filedialog.askdirectory(title=f"Selecciona carpeta de extracción para {os.path.basename(archive_path)}")
        if not dest: return
        ext = os.path.splitext(archive_path)[1].lower()
        try:
            if ext == ".zip":
                shutil.unpack_archive(archive_path, dest, 'zip')
            elif ext == ".rar":
                try:
                    result = subprocess.run(["UnRAR.exe", "x", "-o+", archive_path, dest], capture_output=True, text=True, check=True)
                except (FileNotFoundError, subprocess.CalledProcessError):
                    try:
                        import rarfile
                        rf = rarfile.RarFile(archive_path); rf.extractall(dest)
                    except Exception as e:
                        messagebox.showerror("RAR", f"No se pudo extraer .rar: {e}\nPrueba instalando WinRAR o asegúrate de tener UnRAR.exe en el PATH.")
                        return
            else:
                messagebox.showinfo("Archivo", "Formato no soportado."); return
        except Exception as e:
            messagebox.showerror("Extraer", str(e)); return
        exe_found = None
        for root, dirs, files in os.walk(dest):
            for f in files:
                if f.lower().endswith(('.exe','.msi')):
                    exe_found = os.path.join(root, f); break
            if exe_found: break
        if exe_found:
            if messagebox.askyesno("Ejecutar", f"Se encontró {os.path.basename(exe_found)}. ¿Deseas ejecutarlo?"):
                try: subprocess.Popen([exe_found])
                except Exception as e: messagebox.showerror("Ejecutar", str(e))
        else: messagebox.showinfo("Extraído", "Extraído correctamente, no se encontró ejecutable para lanzar.")

    # ---------------- Run commands in terminals ----------------
    def _set_and_run(self, text):
        self.cmd_var.set(text)
        self.run_command_from_var(text, self.term_text, show_progress=False)

    def run_command_from_var(self, cmd_text, output_widget, show_progress=False):
        raw = cmd_text.strip() if isinstance(cmd_text, str) else ""
        if not raw:
            messagebox.showinfo("Comando", "No hay comandos para ejecutar.")
            return
        lines = [l.strip() for l in raw.splitlines() if l.strip()]
        if not lines: lines = [raw]

        def target():
            for cmd in lines:
                output_widget.insert("end", f"$ {cmd}\n")
                output_widget.see("end")
                try:
                    # ✅ CORREGIDO: Paréntesis cerrado correctamente
                    proc = subprocess.Popen(
                        cmd,
                        shell=True,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        text=True,
                        encoding='utf-8'
                    )
                    for line in proc.stdout:
                        output_widget.insert("end", line)
                        output_widget.see("end")
                    rc = proc.wait()
                    output_widget.insert("end", f"[Proceso finalizado con código {rc}]\n\n")
                    output_widget.see("end")
                except Exception as e:
                    output_widget.insert("end", f"[Error] {e}\n\n")
                    output_widget.see("end")
        threading.Thread(target=target, daemon=True).start()

    def run_command_with_progress(self):
        cmd = self.cmd_var.get().strip()
        if not cmd:
            return

        self.term_text.delete("1.0", "end")
        self.exec_cmd_btn.config(state="disabled")

        def worker():
            try:
                proc = subprocess.Popen(
                    cmd,
                    shell=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding='utf-8'
                )

                def update_progress(percent):
                    blocks = "█" * (percent // 2) + "▒" * (50 - (percent // 2))
                    self.term_text.insert("end", f"  {blocks}  {percent}%\n")
                    self.term_text.see("end")

                # Simular progreso (winget no da % real en upgrade)
                for i in range(0, 101, 2):
                    self.after(0, lambda p=i: update_progress(p))
                    self.after(50)

                for line in proc.stdout:
                    self.term_text.insert("end", line)
                    self.term_text.see("end")

                rc = proc.wait()
                self.term_text.insert("end", f"[Proceso finalizado con código {rc}]\n\n")
                self.term_text.see("end")
            except Exception as e:
                self.term_text.insert("end", f"[Error] {e}\n\n")
                self.term_text.see("end")
            finally:
                self.after(0, lambda: self.exec_cmd_btn.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def open_winstall(self):
        webbrowser.open("https://winstall.app")
        self.repo_status.set("Abriendo Pckgr en winstall.app...")

    # ---------------- Tools check ----------------
    def check_tools(self):
        try:
            if shutil.which("winget") is None:
                self.term_text.insert("end","[Aviso] winget no encontrado. Algunas funciones pueden no funcionar.\n")
        except Exception:
            pass
        try:
            if shutil.which("choco") is None:
                self.term_text.insert("end","[Aviso] chocolatey (choco) no encontrado.\n")
        except Exception:
            pass

    # ---------------- Repo-Software: Winget Backup ----------------
    def list_installed_software(self):
        self.repo_status.set("Obteniendo lista de software instalado...")
        self.software_text.delete("1.0", "end")
        self.software_text.insert("1.0", "Ejecutando 'winget list'...\n\n")

        def worker():
            try:
                proc = subprocess.Popen(
                    ["winget", "list"],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding='utf-8'
                )
                output = ""
                for line in proc.stdout:
                    output += line
                proc.wait()

                if proc.returncode == 0:
                    self.after(0, lambda: self.software_text.delete("1.0", "end"))
                    self.after(0, lambda: self.software_text.insert("1.0", output))
                    self.after(0, lambda: self.repo_status.set("Software listado correctamente."))
                else:
                    self.after(0, lambda: self.software_text.insert("end", "\n[Error al ejecutar winget list]"))
                    self.after(0, lambda: self.repo_status.set("Error al listar software."))
            except Exception as e:
                self.after(0, lambda: self.software_text.insert("end", f"\n[Excepción: {str(e)}]"))
                self.after(0, lambda: self.repo_status.set(f"Error: {str(e)}"))

        threading.Thread(target=worker, daemon=True).start()

    def extract_package_data(self):
        """Extrae datos completos: Nombre, Id, Versión, Disponible, Origen."""
        content = self.software_text.get("1.0", "end").strip()
        if not content or "Nombre" not in content:
            return None

        lines = content.splitlines()
        packages = []
        header_found = False
        col_positions = {}

        for line in lines:
            if "Nombre" in line and "Id" in line and "Versión" in line:
                header_found = True
                col_positions = {
                    "nombre_start": line.find("Nombre"),
                    "id_start": line.find("Id"),
                    "version_start": line.find("Versión"),
                    "disponible_start": line.find("Disponible"),
                    "origen_start": line.find("Origen")
                }
                if any(pos == -1 for pos in col_positions.values()):
                    continue
                continue

            if not header_found or not line.strip() or line.startswith("-"):
                continue

            nombre = line[col_positions["nombre_start"]:col_positions["id_start"]].strip()
            pkg_id = line[col_positions["id_start"]:col_positions["version_start"]].strip()
            version = line[col_positions["version_start"]:col_positions["disponible_start"]].strip() if col_positions["disponible_start"] != -1 else ""
            disponible = line[col_positions["disponible_start"]:col_positions["origen_start"]].strip() if col_positions["origen_start"] != -1 else ""
            origen = line[col_positions["origen_start"]:].strip() if col_positions["origen_start"] != -1 else ""

            if pkg_id and pkg_id != "Id":
                packages.append({
                    "nombre": nombre,
                    "id": pkg_id,
                    "version": version,
                    "disponible": disponible,
                    "origen": origen
                })

        return packages

    def generate_full_backup(self):
        packages = self.extract_package_data()
        if packages is None:
            messagebox.showwarning("Advertencia", "Primero debes listar el software instalado.")
            return
        if not packages:
            messagebox.showwarning("Advertencia", "No se encontraron paquetes válidos.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"winget-backup-{timestamp}"

        dest_folder = filedialog.askdirectory(title="Selecciona carpeta de destino para el backup")
        if not dest_folder:
            return

        bat_path = os.path.join(dest_folder, f"{base_name}.bat")
        json_path = os.path.join(dest_folder, f"{base_name}.json")
        zip_path = os.path.join(dest_folder, f"{base_name}.zip")
        excel_path = os.path.join(dest_folder, f"{base_name}.xlsx")

        # Generar .bat
        try:
            with open(bat_path, "w", encoding="utf-8") as f:
                f.write("@echo off\n")
                f.write(f"REM Backup generado por CSM v.2.0 el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("REM Compatible con Pckgr / Intune\n")
                f.write("echo Iniciando restauración...\n\n")
                for pkg in packages:
                    f.write(f'winget install -e --id "{pkg["id"]}"\n')
                f.write('\necho ¡Restauración completada!\npause\n')
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el archivo .bat:\n{e}")
            return

        # Generar .json
        try:
            data = {"apps": [{"id": pkg["id"]} for pkg in packages]}
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el archivo .json:\n{e}")
            return

        # Generar Excel
        if EXCEL_AVAILABLE:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Software Instalado"

                headers = ["Nombre", "Id", "Versión", "Disponible", "Origen"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                for row_num, pkg in enumerate(packages, 2):
                    ws.cell(row=row_num, column=1, value=pkg["nombre"])
                    ws.cell(row=row_num, column=2, value=pkg["id"])
                    ws.cell(row=row_num, column=3, value=pkg["version"])
                    ws.cell(row=row_num, column=4, value=pkg["disponible"])
                    ws.cell(row=row_num, column=5, value=pkg["origen"])

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = min(adjusted_width, 50)

                wb.save(excel_path)
            except Exception as e:
                messagebox.showwarning("Advertencia Excel", f"No se pudo crear el archivo Excel:\n{e}")

        # Comprimir en .zip
        try:
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(bat_path, os.path.basename(bat_path))
                zipf.write(json_path, os.path.basename(json_path))
                if EXCEL_AVAILABLE and os.path.exists(excel_path):
                    zipf.write(excel_path, os.path.basename(excel_path))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el archivo .zip:\n{e}")
            return

        msg = f"✅ Archivos generados en:\n{dest_folder}\n\n"
        msg += f" - {os.path.basename(bat_path)}\n"
        msg += f" - {os.path.basename(json_path)}\n"
        if EXCEL_AVAILABLE:
            msg += f" - {os.path.basename(excel_path)}\n"
        msg += f" - {os.path.basename(zip_path)}\n\n"
        msg += "El archivo .json es compatible con Pckgr para despliegues en Intune."

        messagebox.showinfo("Backup Completo", msg)
        self.repo_status.set(f"Backup completo generado: {zip_path}")

    # ---------------- Barra de progreso gráfica ----------------
    def run_command_with_progress(self):
        cmd = self.cmd_var.get().strip()
        if not cmd:
            return

        self.term_text.delete("1.0", "end")
        self.exec_cmd_btn.config(state="disabled")

        def worker():
            try:
                proc = subprocess.Popen(
                    cmd,
                    shell=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding='utf-8'
                )

                def update_progress(percent):
                    blocks = "█" * (percent // 2) + "▒" * (50 - (percent // 2))
                    self.term_text.insert("end", f"  {blocks}  {percent}%\n")
                    self.term_text.see("end")

                # Simular progreso (winget no da % real en upgrade)
                for i in range(0, 101, 2):
                    self.after(0, lambda p=i: update_progress(p))
                    self.after(50)

                for line in proc.stdout:
                    self.term_text.insert("end", line)
                    self.term_text.see("end")

                rc = proc.wait()
                self.term_text.insert("end", f"[Proceso finalizado con código {rc}]\n\n")
                self.term_text.see("end")
            except Exception as e:
                self.term_text.insert("end", f"[Error] {e}\n\n")
                self.term_text.see("end")
            finally:
                self.after(0, lambda: self.exec_cmd_btn.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def open_winstall(self):
        webbrowser.open("https://winstall.app")
        self.repo_status.set("Abriendo Pckgr en winstall.app...")

    # ---------------- Tools check ----------------
    def check_tools(self):
        try:
            if shutil.which("winget") is None:
                self.term_text.insert("end","[Aviso] winget no encontrado. Algunas funciones pueden no funcionar.\n")
        except Exception:
            pass
        try:
            if shutil.which("choco") is None:
                self.term_text.insert("end","[Aviso] chocolatey (choco) no encontrado.\n")
        except Exception:
            pass

    # ---------------- Repo-Software: Winget Backup ----------------
    def list_installed_software(self):
        self.repo_status.set("Obteniendo lista de software instalado...")
        self.software_text.delete("1.0", "end")
        self.software_text.insert("1.0", "Ejecutando 'winget list'...\n\n")

        def worker():
            try:
                proc = subprocess.Popen(
                    ["winget", "list"],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding='utf-8'
                )
                output = ""
                for line in proc.stdout:
                    output += line
                proc.wait()

                if proc.returncode == 0:
                    self.after(0, lambda: self.software_text.delete("1.0", "end"))
                    self.after(0, lambda: self.software_text.insert("1.0", output))
                    self.after(0, lambda: self.repo_status.set("Software listado correctamente."))
                else:
                    self.after(0, lambda: self.software_text.insert("end", "\n[Error al ejecutar winget list]"))
                    self.after(0, lambda: self.repo_status.set("Error al listar software."))
            except Exception as e:
                self.after(0, lambda: self.software_text.insert("end", f"\n[Excepción: {str(e)}]"))
                self.after(0, lambda: self.repo_status.set(f"Error: {str(e)}"))

        threading.Thread(target=worker, daemon=True).start()

    def extract_package_data(self):
        """Extrae datos completos: Nombre, Id, Versión, Disponible, Origen."""
        content = self.software_text.get("1.0", "end").strip()
        if not content or "Nombre" not in content:
            return None

        lines = content.splitlines()
        packages = []
        header_found = False
        col_positions = {}

        for line in lines:
            if "Nombre" in line and "Id" in line and "Versión" in line:
                header_found = True
                col_positions = {
                    "nombre_start": line.find("Nombre"),
                    "id_start": line.find("Id"),
                    "version_start": line.find("Versión"),
                    "disponible_start": line.find("Disponible"),
                    "origen_start": line.find("Origen")
                }
                if any(pos == -1 for pos in col_positions.values()):
                    continue
                continue

            if not header_found or not line.strip() or line.startswith("-"):
                continue

            nombre = line[col_positions["nombre_start"]:col_positions["id_start"]].strip()
            pkg_id = line[col_positions["id_start"]:col_positions["version_start"]].strip()
            version = line[col_positions["version_start"]:col_positions["disponible_start"]].strip() if col_positions["disponible_start"] != -1 else ""
            disponible = line[col_positions["disponible_start"]:col_positions["origen_start"]].strip() if col_positions["origen_start"] != -1 else ""
            origen = line[col_positions["origen_start"]:].strip() if col_positions["origen_start"] != -1 else ""

            if pkg_id and pkg_id != "Id":
                packages.append({
                    "nombre": nombre,
                    "id": pkg_id,
                    "version": version,
                    "disponible": disponible,
                    "origen": origen
                })

        return packages

    def generate_full_backup(self):
        packages = self.extract_package_data()
        if packages is None:
            messagebox.showwarning("Advertencia", "Primero debes listar el software instalado.")
            return
        if not packages:
            messagebox.showwarning("Advertencia", "No se encontraron paquetes válidos.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"winget-backup-{timestamp}"

        dest_folder = filedialog.askdirectory(title="Selecciona carpeta de destino para el backup")
        if not dest_folder:
            return

        bat_path = os.path.join(dest_folder, f"{base_name}.bat")
        json_path = os.path.join(dest_folder, f"{base_name}.json")
        zip_path = os.path.join(dest_folder, f"{base_name}.zip")
        excel_path = os.path.join(dest_folder, f"{base_name}.xlsx")

        # Generar .bat
        try:
            with open(bat_path, "w", encoding="utf-8") as f:
                f.write("@echo off\n")
                f.write(f"REM Backup generado por CSM v.2.0 el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("REM Compatible con Pckgr / Intune\n")
                f.write("echo Iniciando restauración...\n\n")
                for pkg in packages:
                    f.write(f'winget install -e --id "{pkg["id"]}"\n')
                f.write('\necho ¡Restauración completada!\npause\n')
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el archivo .bat:\n{e}")
            return

        # Generar .json
        try:
            data = {"apps": [{"id": pkg["id"]} for pkg in packages]}
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el archivo .json:\n{e}")
            return

        # Generar Excel
        if EXCEL_AVAILABLE:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Software Instalado"

                headers = ["Nombre", "Id", "Versión", "Disponible", "Origen"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                for row_num, pkg in enumerate(packages, 2):
                    ws.cell(row=row_num, column=1, value=pkg["nombre"])
                    ws.cell(row=row_num, column=2, value=pkg["id"])
                    ws.cell(row=row_num, column=3, value=pkg["version"])
                    ws.cell(row=row_num, column=4, value=pkg["disponible"])
                    ws.cell(row=row_num, column=5, value=pkg["origen"])

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = min(adjusted_width, 50)

                wb.save(excel_path)
            except Exception as e:
                messagebox.showwarning("Advertencia Excel", f"No se pudo crear el archivo Excel:\n{e}")

        # Comprimir en .zip
        try:
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(bat_path, os.path.basename(bat_path))
                zipf.write(json_path, os.path.basename(json_path))
                if EXCEL_AVAILABLE and os.path.exists(excel_path):
                    zipf.write(excel_path, os.path.basename(excel_path))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el archivo .zip:\n{e}")
            return

        msg = f"✅ Archivos generados en:\n{dest_folder}\n\n"
        msg += f" - {os.path.basename(bat_path)}\n"
        msg += f" - {os.path.basename(json_path)}\n"
        if EXCEL_AVAILABLE:
            msg += f" - {os.path.basename(excel_path)}\n"
        msg += f" - {os.path.basename(zip_path)}\n\n"
        msg += "El archivo .json es compatible con Pckgr para despliegues en Intune."

        messagebox.showinfo("Backup Completo", msg)
        self.repo_status.set(f"Backup completo generado: {zip_path}")

    # ---------------- Barra de progreso gráfica ----------------
    def run_command_with_progress(self):
        cmd = self.cmd_var.get().strip()
        if not cmd:
            return

        self.term_text.delete("1.0", "end")
        self.exec_cmd_btn.config(state="disabled")

        def worker():
            try:
                proc = subprocess.Popen(
                    cmd,
                    shell=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding='utf-8'
                )

                def update_progress(percent):
                    blocks = "█" * (percent // 2) + "▒" * (50 - (percent // 2))
                    self.term_text.insert("end", f"  {blocks}  {percent}%\n")
                    self.term_text.see("end")

                # Simular progreso (winget no da % real en upgrade)
                for i in range(0, 101, 2):
                    self.after(0, lambda p=i: update_progress(p))
                    self.after(50)

                for line in proc.stdout:
                    self.term_text.insert("end", line)
                    self.term_text.see("end")

                rc = proc.wait()
                self.term_text.insert("end", f"[Proceso finalizado con código {rc}]\n\n")
                self.term_text.see("end")
            except Exception as e:
                self.term_text.insert("end", f"[Error] {e}\n\n")
                self.term_text.see("end")
            finally:
                self.after(0, lambda: self.exec_cmd_btn.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def open_winstall(self):
        webbrowser.open("https://winstall.app")
        self.repo_status.set("Abriendo Pckgr en winstall.app...")

# --- PUNTO DE ENTRADA ---
if __name__ == "__main__":
    app = InstallerApp()
    app.mainloop()